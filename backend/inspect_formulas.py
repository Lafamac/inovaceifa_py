import openpyxl
import sys

wb = openpyxl.load_workbook('planilha.xlsx', read_only=True)

with open('inspect_formulas_output.txt', 'w', encoding='utf-8') as f:
    def log_rows(sheet_name, rows):
        f.write(f"\n=========================================\n")
        f.write(f"SHEET: {sheet_name}\n")
        f.write(f"=========================================\n")
        if sheet_name not in wb.sheetnames:
            f.write("Not found!\n")
            return
        sheet = wb[sheet_name]
        for r in rows:
            try:
                row = next(sheet.iter_rows(min_row=r, max_row=r, max_col=35, values_only=True))
                row_vals = [str(x) if x is not None else "" for x in row]
                f.write(f"Row {r}: {row_vals}\n")
            except Exception as e:
                f.write(f"Error Row {r}: {e}\n")

    log_rows('Aba Rateios', range(1, 10))
    log_rows('Custos_Talhão', range(1, 26))
